from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from postavleno_bot.integrations.wildberries import WBStockItem
from postavleno_bot.services.wb_export import (
    build_export_dataframe,
    build_export_filename,
    build_export_xlsx,
)
from postavleno_bot.utils.excel import dataframe_to_xlsx_bytes, save_dataframe_to_xlsx


def _item(
    warehouse: str,
    *,
    article: str,
    nm: int,
    quantity: int,
    price: int | None = None,
    discount: int | None = None,
) -> WBStockItem:
    return WBStockItem(
        lastChangeDate=datetime(2024, 1, 1, 12, 0),
        warehouseName=warehouse,
        supplierArticle=article,
        nmId=nm,
        barcode=f"bc-{nm}",
        quantity=quantity,
        inWayToClient=0,
        inWayFromClient=0,
        quantityFull=quantity,
        category="Категория",
        subject="Предмет",
        brand="Бренд",
        techSize="M",
        price=price,
        discount=discount,
        scCode=None,
    )


def test_export_xlsx_headers() -> None:
    items = [
        _item("Склад B", article="B", nm=2, quantity=5, price=1000, discount=10),
        _item("Склад A", article="A", nm=1, quantity=7, price=1500, discount=5),
    ]

    data = build_export_xlsx(items, sheet_name="FootballShop")
    df = pd.read_excel(io.BytesIO(data))

    assert list(df.columns) == [
        "Склад",
        "Артикул",
        "nmId",
        "Штрихкод",
        "Кол-во",
        "Категория",
        "Предмет",
        "Бренд",
        "Размер",
        "Цена",
        "Скидка",
    ]

    # Sorted by warehouse asc, quantity desc.
    first_row = df.iloc[0].to_dict()
    assert first_row["Склад"] == "Склад A"
    assert first_row["Артикул"] == "A"
    assert first_row["nmId"] == 1
    assert first_row["Кол-во"] == 7


def test_excel_roundtrip() -> None:
    frame = pd.DataFrame(
        {
            "Колонка A": [1, 2, 3],
            "Колонка B": ["x", "y", "z"],
        }
    )

    payload = dataframe_to_xlsx_bytes(frame, sheet_name="Sheet")
    restored = pd.read_excel(io.BytesIO(payload))

    assert restored.equals(frame)


def test_filename_all_vs_single() -> None:
    moment = datetime(2024, 1, 2, 15, 30)
    assert build_export_filename("ALL", None, moment) == "wb_ostatki_ALL_20240102_1530.xlsx"
    assert (
        build_export_filename("wh", "Склад Москва", moment)
        == "wb_ostatki_Склад_Москва_20240102_1530.xlsx"
    )


def test_export_dataframe_structure() -> None:
    items = [
        _item("Склад A", article="A", nm=1, quantity=5),
        _item("Склад A", article="B", nm=2, quantity=3),
    ]

    frame = build_export_dataframe(items)

    assert list(frame.columns) == [
        "Склад",
        "Артикул",
        "nmId",
        "Штрихкод",
        "Кол-во",
        "Категория",
        "Предмет",
        "Бренд",
        "Размер",
        "Цена",
        "Скидка",
    ]
    assert frame.iloc[0]["Склад"] == "Склад A"


def test_excel_xlsxwriter_opens(tmp_path) -> None:
    frame = pd.DataFrame({"A": [1, 2], "B": ["x", "y"]})
    path = save_dataframe_to_xlsx(frame, path=tmp_path / "sample.xlsx", sheet_name="Data")

    workbook = load_workbook(path)
    sheet = workbook["Data"]
    values = list(sheet.iter_rows(min_row=2, max_row=3, values_only=True))
    assert values == [(1, "x"), (2, "y")]
